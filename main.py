import csv
import io
import threading
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.request import Request, urlopen

from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import FileResponse, JSONResponse
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
app = FastAPI(title="Certification Program Dashboard")

WORKBOOK_URL = (
    "https://docs.google.com/spreadsheets/d/e/2PACX-1vRdGlnE53K-er93AZC_2Na83aYvoGUNB0kHoAXRKHRXvNWYwufokS8GGdpf2qrCa0zUrr25uSBXUMVe/pub?output=xlsx"
)

SHEET_TABS: Dict[str, List[str]] = {
    "fullPaymentCohort": [
        "Cert Program Full Payment Cohort",
        "Cert Program Full Payment Cohor",
    ],
    "fullPaymentMonthly": ["Cert Program Full Payment Month"],
    "tlCohort": ["Cert TL Wise Cohort Full"],
    "tlMonthly": ["Cert TL Wise Monthly Full"],
    "gmCohort": ["Cert GM Wise Cohort Full"],
    "gmMonthly": ["Cert GM Wise Monthly Full"],
    "bdaCohort": ["Cert BDA Wise Cohort Full"],
    "bdaMonthly": ["Cert BDA Wise Monthly Full"],
}

_workbook_cache: Optional[object] = None
_workbook_cache_at: float = 0.0
_sheet_csv_cache: Dict[str, Tuple[str, str]] = {}
_workbook_lock = threading.Lock()
CACHE_SECONDS = 300


def _normalize_name(name: str) -> str:
    return (name or "").strip().lower()


def _resolve_sheet_name(sheet_names: List[str], candidates: List[str]) -> Optional[str]:
    by_norm = {_normalize_name(name): name for name in sheet_names}
    for candidate in candidates:
        match = by_norm.get(_normalize_name(candidate))
        if match:
            return match
    for candidate in candidates:
        needle = _normalize_name(candidate)
        for name in sheet_names:
            norm = _normalize_name(name)
            if norm.startswith(needle) or needle.startswith(norm):
                return name
    return None


def _load_workbook():
    global _workbook_cache, _workbook_cache_at, _sheet_csv_cache
    now = time.time()
    with _workbook_lock:
        if _workbook_cache is not None and now - _workbook_cache_at < CACHE_SECONDS:
            return _workbook_cache

        request = Request(WORKBOOK_URL, headers={"User-Agent": "CertDashboard/1.0"})
        with urlopen(request, timeout=120) as response:
            payload = response.read()

        _workbook_cache = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        _workbook_cache_at = now
        _sheet_csv_cache = {}
        return _workbook_cache


def _worksheet_to_csv(worksheet) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    for row in worksheet.iter_rows(values_only=True):
        writer.writerow(["" if cell is None else cell for cell in row])
    return buffer.getvalue()


def _get_sheet_csv(sheet: str) -> Tuple[str, str]:
    cached = _sheet_csv_cache.get(sheet)
    if cached:
        return cached

    candidates = SHEET_TABS[sheet]
    workbook = _load_workbook()
    tab_name = _resolve_sheet_name(workbook.sheetnames, candidates)
    if not tab_name:
        raise HTTPException(status_code=404, detail=f"Workbook tab not found for {sheet}")

    csv_text = _worksheet_to_csv(workbook[tab_name])
    _sheet_csv_cache[sheet] = (tab_name, csv_text)
    return tab_name, csv_text


@app.get("/api/sheets")
def api_sheets(sheet: str = Query(..., description="Dataset key")):
    if sheet not in SHEET_TABS:
        raise HTTPException(status_code=400, detail="Invalid or missing sheet key")

    try:
        tab_name, csv_text = _get_sheet_csv(sheet)
        return JSONResponse({"sheet": sheet, "tab": tab_name, "csv": csv_text})
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


NO_CACHE_HEADERS = {
    "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
    "Pragma": "no-cache",
}


@app.get("/")
def read_dashboard() -> FileResponse:
    return FileResponse(BASE_DIR / "index.html", headers=NO_CACHE_HEADERS)


@app.get("/{file_path:path}")
def read_static_file(file_path: str):
    requested = (BASE_DIR / file_path).resolve()
    if requested.is_file() and requested.parent == BASE_DIR:
        return FileResponse(requested, headers=NO_CACHE_HEADERS)
    return FileResponse(BASE_DIR / "index.html", headers=NO_CACHE_HEADERS)
